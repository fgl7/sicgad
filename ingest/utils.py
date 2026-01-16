from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Iterable, List, Tuple

from io import TextIOWrapper, StringIO
import csv

import openpyxl

from .models import DatasetInstance, PublishedDataPoint
from schemas.models import ColumnDef


@dataclass
class ParsedRow:
  row_index: int
  values: List[str]


MONTH_ALIASES = {
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "febrero": 2,
    "mar": 3,
    "marzo": 3,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "junio": 6,
    "jul": 7,
    "julio": 7,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "setiembre": 9,
    "oct": 10,
    "octubre": 10,
    "nov": 11,
    "noviembre": 11,
    "dic": 12,
    "diciembre": 12,
}

MONTH_NAMES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def read_uploaded_file(uploaded_file) -> Tuple[List[str], List[List[object]]]:
    """
    Lee un archivo subido (CSV/Excel) y devuelve encabezado y filas (valores crudos).
    A diferencia de `_read_instance_file`, esta función no depende de DatasetInstance.
    """

    if not uploaded_file:
        return [], []

    name = getattr(uploaded_file, "name", "") or ""
    ext = name.lower().rsplit(".", 1)[-1] if "." in name else ""

    header: List[str] = []
    rows: List[List[object]] = []
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    if ext in ("xlsx", "xlsm", "xltx", "xltm"):
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            values = list(row)
            if i == 0:
                header = ["" if v is None else str(v) for v in values]
            else:
                rows.append(values)
    else:
        raw_bytes = uploaded_file.read()
        if isinstance(raw_bytes, bytes):
            text = raw_bytes.decode("utf-8", errors="ignore")
        else:
            text = str(raw_bytes)
        reader = csv.reader(StringIO(text))
        for i, row in enumerate(reader):
            if i == 0:
                header = row
            else:
                rows.append(list(row))
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    return header, rows


def parse_date_cell(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    # Intentos comunes
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            parsed_dt = datetime.strptime(text, fmt)
            return parsed_dt.date()
        except ValueError:
            continue
    try:
        parsed = date.fromisoformat(text[:10])
        return parsed
    except Exception:
        return None


def month_number_from_label(label: str | None) -> int | None:
    if not label:
        return None
    text = str(label).strip().lower()
    if not text:
        return None
    if text in MONTH_ALIASES:
        return MONTH_ALIASES[text]
    key = text[:3]
    return MONTH_ALIASES.get(key)


def month_label(month: int | None) -> str:
    if not month:
        return ""
    return MONTH_NAMES.get(month, "")


def month_columns_for_dataset(dataset) -> dict[str, int]:
    if not dataset:
        return {}
    month_columns: dict[str, int] = {}
    for column in dataset.columns.filter(is_active=True):
        month = month_number_from_label(column.name) or month_number_from_label(column.label)
        if month:
            month_columns[column.name] = month
    return month_columns

def _read_instance_file(instance: DatasetInstance) -> Tuple[List[str], List[ParsedRow]]:
    """
    Lee el archivo bruto de una instancia y devuelve encabezado y filas.
    No aplica ningún tipo de validación de negocio; solo lectura.
    """

    header: List[str] = []
    rows: List[ParsedRow] = []

    if not instance.raw_file:
        return header, rows

    file_field = instance.raw_file
    name = file_field.name or ""
    ext = name.lower().rsplit(".", 1)[-1] if "." in name else ""

    if ext in ("xlsx", "xlsm", "xltx", "xltm"):
        with file_field.open("rb") as fh:
            wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                values = ["" if v is None else str(v) for v in row]
                if i == 0:
                    header = values
                else:
                    rows.append(ParsedRow(row_index=i, values=values))
    else:
        with file_field.open("rb") as fh:
            wrapper = TextIOWrapper(fh, encoding="utf-8", errors="ignore")
            reader = csv.reader(wrapper)
            for i, row in enumerate(reader):
                values = row
                if i == 0:
                    header = values
                else:
                    rows.append(ParsedRow(row_index=i, values=values))

    return header, rows


def _parse_value(raw: str, column: ColumnDef):
    if raw is None:
        return None, None, None, None

    raw = str(raw).strip()
    if raw == "":
        return None, None, None, None

    dt = column.data_type

    if dt in ("INTEGER", "FLOAT"):
        try:
            return float(raw), None, None, None
        except ValueError:
            return None, raw, None, None

    if dt == "DATE":
        if isinstance(raw, (date, datetime)):
            return None, None, raw.date() if isinstance(raw, datetime) else raw, None
        try:
            parsed = date.fromisoformat(raw)
            return None, None, parsed, None
        except Exception:
            return None, raw, None, None

    if dt == "BOOLEAN":
        lowered = raw.lower()
        if lowered in ("1", "true", "t", "yes", "y", "si", "sí"):
            return None, None, None, True
        if lowered in ("0", "false", "f", "no", "n"):
            return None, None, None, False
        return None, raw, None, None

    # STRING / CHOICE / otros
    return None, raw, None, None


def materialize_instance(instance: DatasetInstance) -> int:
    """
    Convierte el archivo bruto de una instancia publicada en filas de PublishedDataPoint.
    Si ya existían puntos para la instancia, se eliminan primero.
    Devuelve la cantidad de puntos creados.
    """

    # Limpieza previa
    PublishedDataPoint.objects.filter(instance=instance).delete()

    header, rows = _read_instance_file(instance)
    if not header or not rows:
        return 0

    # Mapeo encabezado -> ColumnDef
    dataset = instance.dataset_type
    columns = list(dataset.columns.all())

    header_map: dict[str, ColumnDef] = {}
    for col in columns:
        # Por definición, el encabezado oficial es el "name" de la columna.
        base = (col.name or "").strip()
        if base:
            header_map[base.lower()] = col
        # Aceptamos también la etiqueta como encabezado, por compatibilidad hacia atrás.
        alt = (col.label or "").strip()
        if alt:
            header_map.setdefault(alt.lower(), col)

    index_to_column: List[ColumnDef | None] = []
    for name in header:
        col = header_map.get((name or "").strip().lower())
        index_to_column.append(col)

    points: list[PublishedDataPoint] = []
    for row in rows:
        for idx, raw in enumerate(row.values):
            column = index_to_column[idx] if idx < len(index_to_column) else None
            if not column:
                continue

            numeric_value, text_value, date_value, bool_value = _parse_value(raw, column)
            if (
                numeric_value is None
                and text_value in (None, "")
                and date_value is None
                and bool_value is None
            ):
                continue

            points.append(
                PublishedDataPoint(
                    instance=instance,
                    column=column,
                    row_index=row.row_index,
                    numeric_value=numeric_value,
                    text_value=text_value or "",
                    date_value=date_value,
                    bool_value=bool_value,
                )
            )

    if points:
        PublishedDataPoint.objects.bulk_create(points, batch_size=1000)

    return len(points)
