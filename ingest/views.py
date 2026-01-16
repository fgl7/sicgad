import calendar
from datetime import date, datetime

from django.contrib import messages
from django.http import Http404, HttpResponse
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.db.models import Q, Sum, Min, Max, Count
from django.forms import formset_factory
from django.core.files.base import ContentFile
from io import BytesIO, TextIOWrapper, StringIO
import csv
import openpyxl

from accounts.models import Membership
from plants.models import Plant
from projects.models import Project
from schemas.models import DatasetType
from schemas.services import previous_month_range
from audit.utils import record_action
from .forms import (
    DatasetInstanceUploadForm,
    DatasetInstanceEditForm,
    ManualDatasetForm,
    build_manual_row_form,
    CertificationJustificationForm,
    HistoricalDatasetUploadForm,
)
from .models import (
    DatasetInstance,
    PublishedDataPoint,
    DatasetChangeRequest,
    DatasetChangeAttachment,
    HistoricalImportBatch,
)
from .utils import (
    _read_instance_file,
    month_columns_for_dataset,
    month_label,
    month_number_from_label,
    parse_date_cell,
    read_uploaded_file,
)


def _has_one_time_instance(dataset: DatasetType | None) -> bool:
    if not dataset or not dataset.is_one_time:
        return False
    qs = DatasetInstance.objects.filter(dataset_type=dataset)
    if dataset.plant_id:
        qs = qs.filter(plant_id=dataset.plant_id)
    elif dataset.project_id:
        qs = qs.filter(project_id=dataset.project_id)
    return qs.exists()


def _is_weekly_month_locked_dataset(dataset: DatasetType | None) -> bool:
    if not dataset or dataset.validation_frequency != DatasetType.WEEKLY:
        return False
    if dataset.is_one_time:
        return False
    name = (dataset.name or "").lower()
    slug = (dataset.slug or "").lower()
    return "curva_s_ejecutado" in name or "curva_s_ejecutado" in slug or "curva-s-ejecutado" in slug


def _normalize_month_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return text.lower()


def _values_match(left, right) -> bool:
    if left is None and right is None:
        return True
    if left is None or right is None:
        return False
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(left - right) < 1e-6
    return str(left).strip().lower() == str(right).strip().lower()


def _month_values_from_rows(header, rows, dataset_months: set[int]) -> dict[int, object]:
    if not header:
        return {}
    header_months = []
    for idx, name in enumerate(header):
        month = month_number_from_label(name)
        if month and month in dataset_months:
            header_months.append((idx, month))

    values: dict[int, object] = {}
    for row in rows:
        row_values = row.values if hasattr(row, "values") else row
        for idx, month in header_months:
            if idx >= len(row_values):
                continue
            parsed = _normalize_month_value(row_values[idx])
            if parsed is None:
                continue
            values[month] = parsed
    return values


def _previous_month_values(dataset: DatasetType, period: date | None) -> dict[int, object]:
    if not period:
        return {}
    qs = DatasetInstance.objects.filter(dataset_type=dataset, period__lt=period)
    if dataset.plant_id:
        qs = qs.filter(plant_id=dataset.plant_id)
    elif dataset.project_id:
        qs = qs.filter(project_id=dataset.project_id)
    prev_instance = qs.order_by("-period", "-created_at").first()
    if not prev_instance:
        return {}

    header, rows = _read_instance_file(prev_instance)
    dataset_months = set(month_columns_for_dataset(dataset).values())
    return _month_values_from_rows(header, rows, dataset_months)


def _validate_weekly_month_rows(dataset: DatasetType, period: date | None, rows: list[dict]) -> str | None:
    if not period or not _is_weekly_month_locked_dataset(dataset):
        return None
    month_columns = month_columns_for_dataset(dataset)
    if not month_columns:
        return None
    allowed_month = period.month
    dataset_months = set(month_columns.values())
    previous_values = _previous_month_values(dataset, period)
    current_values: dict[int, object] = {}
    for row in rows:
        for col_name, month in month_columns.items():
            if col_name not in row:
                continue
            parsed = _normalize_month_value(row.get(col_name))
            if parsed is None:
                continue
            current_values[month] = parsed

    mismatched_months = []
    for month in sorted(dataset_months):
        if month == allowed_month:
            continue
        if not _values_match(previous_values.get(month), current_values.get(month)):
            mismatched_months.append(month_label(month) or str(month))

    if mismatched_months:
        allowed_label = month_label(allowed_month) or str(allowed_month)
        blocked_label = ", ".join(mismatched_months)
        return (
            f"Solo puedes actualizar el mes de {allowed_label}. "
            f"No modifiques los meses: {blocked_label}."
        )
    return None


def _validate_weekly_month_file(
    dataset: DatasetType, period: date | None, uploaded_file
) -> str | None:
    if not period or not _is_weekly_month_locked_dataset(dataset):
        return None
    month_columns = month_columns_for_dataset(dataset)
    if not month_columns:
        return None

    header, rows = read_uploaded_file(uploaded_file)
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    if not header:
        return "El archivo no tiene encabezado."

    dataset_months = set(month_columns.values())
    header_months = []
    for idx, name in enumerate(header):
        month = month_number_from_label(name)
        if month and month in dataset_months:
            header_months.append((idx, month, name))

    allowed_month = period.month
    if allowed_month not in dataset_months:
        return None
    if not any(month == allowed_month for _, month, _ in header_months):
        allowed_label = month_label(allowed_month) or str(allowed_month)
        return f"El archivo no incluye la columna del mes de {allowed_label}."

    current_values = _month_values_from_rows(header, rows, dataset_months)
    previous_values = _previous_month_values(dataset, period)

    mismatched_months = []
    for month in sorted(dataset_months):
        if month == allowed_month:
            continue
        if not _values_match(previous_values.get(month), current_values.get(month)):
            mismatched_months.append(month_label(month) or str(month))

    if mismatched_months:
        allowed_label = month_label(allowed_month) or str(allowed_month)
        blocked_label = ", ".join(mismatched_months)
        return (
            f"Solo puedes actualizar el mes de {allowed_label}. "
            f"No modifiques los meses: {blocked_label}."
        )
    return None


def _apply_weekly_month_lock_to_formset(
    row_formset,
    dataset: DatasetType,
    period: date | None,
    previous_values: dict[int, object] | None = None,
) -> bool:
    if not row_formset or not period or not _is_weekly_month_locked_dataset(dataset):
        return False
    month_columns = month_columns_for_dataset(dataset)
    if not month_columns:
        return False
    allowed_month = period.month
    for form in row_formset:
        for col_name, month in month_columns.items():
            if month == allowed_month or col_name not in form.fields:
                continue
            field = form.fields[col_name]
            if previous_values and month in previous_values:
                form.initial[col_name] = previous_values[month]
                field.initial = previous_values[month]
            field.required = False
            field.disabled = True
            field.widget.attrs["disabled"] = True
    return True


def upload(request):
    user = request.user
    if user.is_authenticated and (
        user.is_superuser
        or Membership.objects.filter(user=user, role="ADMIN", is_active=True).exists()
    ):
        messages.info(
            request,
            "Los administradores no realizan cargas directas; use el historial para revisar y gestionar cargas existentes.",
        )
        return redirect(reverse("ingest:upload_history"))

    loader_plants = None
    loader_projects = None
    if user.is_authenticated:
        loader_memberships = (
            Membership.objects.filter(user=user, role="LOADER", is_active=True)
            .select_related("plant", "project")
        )
        has_global_loader = loader_memberships.filter(
            plant__isnull=True, project__isnull=True
        ).exists()
        if not has_global_loader:
            loader_plants = list(
                loader_memberships.exclude(plant__isnull=True).values_list(
                    "plant_id", flat=True
                )
            )
            loader_projects = list(
                loader_memberships.exclude(project__isnull=True).values_list(
                    "project_id", flat=True
                )
            )

    default_plant = Plant.objects.filter(id=loader_plants[0]).first() if loader_plants else None

    if request.method == "POST":
        form = DatasetInstanceUploadForm(
            request.POST,
            request.FILES,
            loader_plants=loader_plants,
            loader_projects=loader_projects,
        )
        if form.is_valid():
            instance: DatasetInstance = form.save(commit=False)
            upload_error = _validate_weekly_month_file(
                instance.dataset_type,
                instance.period,
                form.cleaned_data.get("raw_file"),
            )
            if upload_error:
                form.add_error("raw_file", upload_error)
            elif _has_one_time_instance(instance.dataset_type):
                messages.error(
                    request,
                    "Este esquema es de carga unica y ya tiene una carga registrada.",
                )
                return redirect(reverse("ingest:upload"))
            else:
                if request.user.is_authenticated:
                    membership = (
                        Membership.objects.filter(user=request.user, is_active=True)
                        .filter(
                            Q(plant=instance.plant)
                            | Q(project=instance.project)
                            | Q(plant__isnull=True, project__isnull=True)
                        )
                        .order_by("role")
                        .first()
                    )
                else:
                    membership = None

                instance.created_by = membership
                if membership and membership.role == "LOADER":
                    # Asegura consistencia: el dataset define la planta/proyecto destino.
                    instance.plant = instance.dataset_type.plant
                    instance.project = instance.dataset_type.project
                instance.state = DatasetInstance.STATE_DRAFT
                instance.row_count = 0
                instance.error_count = 0
                instance.last_error_summary = ""
                instance.save()

                messages.success(
                    request,
                    "Archivo subido correctamente. Revisa tus cargas y env?a el dataset a validaci?n cuando corresponda.",
                )
                record_action(
                    "UPLOAD",
                    request=request,
                    module="Ingest",
                    object_repr=f"{instance.dataset_type.name} | {instance.period}",
                    details=(
                        f"Planta {instance.plant.code}"
                        if instance.plant
                        else f"Proyecto {instance.project.name}"
                    ),
                )
                return redirect(reverse("ingest:upload_history"))
    else:
        form = DatasetInstanceUploadForm(
            loader_plants=loader_plants,
            loader_projects=loader_projects,
        )

    if request.user.is_authenticated:
        instances = (
            DatasetInstance.objects.select_related("dataset_type", "plant", "project")
            .filter(created_by__user=request.user)
            .order_by("-created_at")[:10]
        )
    else:
        instances = DatasetInstance.objects.none()

    return render(request, "ingest/upload.html", {"form": form, "instances": instances})


def dataset_has_data(request):
    user = request.user
    if not user.is_authenticated:
        return JsonResponse(
            {
                "has_data": False,
                "validation_frequency": None,
                "is_certification": False,
            }
        )

    dataset_type_id = request.GET.get("dataset_type")
    try:
        dataset_type_id_int = int(dataset_type_id) if dataset_type_id else None
    except (TypeError, ValueError):
        return JsonResponse(
            {
                "has_data": False,
                "validation_frequency": None,
                "is_certification": False,
            }
        )

    if not dataset_type_id_int:
        return JsonResponse(
            {
                "has_data": False,
                "validation_frequency": None,
                "is_certification": False,
            }
        )

    is_loader = Membership.objects.filter(user=user, role="LOADER", is_active=True).exists()
    if not is_loader:
        return JsonResponse(
            {
                "has_data": False,
                "validation_frequency": None,
                "is_certification": False,
            }
        )

    dataset = (
        DatasetType.objects.select_related("plant", "project")
        .filter(pk=dataset_type_id_int)
        .first()
    )
    if not dataset or (not dataset.plant_id and not dataset.project_id):
        return JsonResponse(
            {
                "has_data": False,
                "validation_frequency": None,
                "is_certification": False,
            }
        )

    # El gating de histórico aplica solo a datasets diarios. Para semanales/mensuales
    # se considera habilitado (no bloquea la UI).
    if dataset.validation_frequency != DatasetType.DAILY:
        return JsonResponse(
            {
                "has_data": True,
                "validation_frequency": dataset.validation_frequency,
                "is_certification": bool(dataset.is_certification),
            }
        )

    has_data = DatasetInstance.objects.filter(
        dataset_type_id=dataset_type_id_int,
    )
    if dataset.plant_id:
        has_data = has_data.filter(plant_id=dataset.plant_id)
    else:
        has_data = has_data.filter(project_id=dataset.project_id)
    has_data = has_data.exists()

    return JsonResponse(
        {
            "has_data": bool(has_data),
            "validation_frequency": dataset.validation_frequency,
            "is_certification": bool(dataset.is_certification),
        }
    )


def upload_historical(request):
    user = request.user
    if user.is_authenticated and (
        user.is_superuser
        or Membership.objects.filter(user=user, role="ADMIN", is_active=True).exists()
    ):
        messages.info(
            request,
            "Los administradores no realizan cargas directas; use el historial para revisar y gestionar cargas existentes.",
        )
        return redirect(reverse("ingest:upload_history"))

    if not user.is_authenticated:
        return redirect("login")

    loader_memberships = (
        Membership.objects.filter(user=user, role="LOADER", is_active=True)
        .select_related("plant")
    )
    has_global_loader = loader_memberships.filter(
        plant__isnull=True, project__isnull=True
    ).exists()
    if has_global_loader:
        loader_plants = list(Plant.objects.values_list("id", flat=True))
    else:
        loader_plants = list(
            loader_memberships.exclude(plant__isnull=True).values_list("plant_id", flat=True)
        )

    if request.method == "POST":
        form = HistoricalDatasetUploadForm(
            request.POST,
            request.FILES,
            loader_plants=loader_plants,
        )
        if form.is_valid():
            dataset_type = form.cleaned_data["dataset_type"]
            plant = form.cleaned_data["plant"]
            date_column_name = (form.cleaned_data["date_column_name"] or "").strip()
            uploaded = form.cleaned_data["raw_file"]

            membership = Membership.objects.filter(
                user=user,
                role="LOADER",
                is_active=True,
                plant=plant,
            ).first()
            if not membership:
                membership = Membership.objects.filter(
                    user=user,
                    role="LOADER",
                    is_active=True,
                    plant__isnull=True,
                    project__isnull=True,
                ).first()

            batch = HistoricalImportBatch.objects.create(
                dataset_type=dataset_type,
                plant=plant,
                created_by=membership,
                date_column_name=date_column_name or "(auto)",
                status=HistoricalImportBatch.STATUS_RUNNING,
            )

            try:
                header, rows = read_uploaded_file(uploaded)
                batch.total_rows = len(rows)

                header_norm = [(h or "").strip().lower() for h in (header or [])]
                if not header_norm:
                    raise ValueError("El archivo no tiene encabezado.")

                # Determinar la columna fecha: si el usuario no la indicó, usamos
                # la primera columna activa de tipo DATE del esquema.
                if not date_column_name:
                    date_col = (
                        dataset_type.columns.filter(is_active=True, data_type="DATE")
                        .order_by("display_order", "name")
                        .first()
                    )
                    if not date_col:
                        raise ValueError(
                            "Este esquema no tiene una columna de tipo DATE. Indica manualmente el encabezado de fecha."
                        )
                    candidates = [date_col.name, date_col.label]
                else:
                    candidates = [date_column_name]

                candidates_norm = [
                    (c or "").strip().lower() for c in candidates if (c or "").strip()
                ]
                date_idx = None
                for cand in candidates_norm:
                    if cand in header_norm:
                        date_idx = header_norm.index(cand)
                        break
                if date_idx is None:
                    raise ValueError(
                        "No se encontró la columna de fecha en el archivo. "
                        "Indica el encabezado exacto o asegúrate de que exista la columna DATE del esquema (por name o label)."
                    )
                batch.date_column_name = header[date_idx] or candidates[0] or "(auto)"

                grouped: dict[date, list[list[object]]] = {}
                for row in rows:
                    cell_value = row[date_idx] if date_idx < len(row) else None
                    period = parse_date_cell(cell_value)
                    if not period:
                        continue
                    grouped.setdefault(period, []).append(row)

                if not grouped:
                    raise ValueError(
                        "No se encontraron filas con fecha válida. Revisa el formato de la columna de fecha."
                    )

                batch.total_days = len(grouped)
                all_periods = sorted(grouped.keys())
                batch.period_start = all_periods[0]
                batch.period_end = all_periods[-1]

                created_count = 0
                updated_count = 0
                skipped_count = 0

                for period, day_rows in grouped.items():
                    instance = DatasetInstance.objects.filter(
                        dataset_type=dataset_type,
                        plant=plant,
                        period=period,
                    ).first()

                    if instance and instance.state != DatasetInstance.STATE_DRAFT:
                        skipped_count += 1
                        continue

                    if not instance:
                        instance = DatasetInstance(
                            dataset_type=dataset_type,
                            plant=plant,
                            period=period,
                        )
                        created_count += 1
                    else:
                        updated_count += 1

                    instance.created_by = membership
                    instance.historical_batch = batch
                    instance.state = DatasetInstance.STATE_DRAFT
                    instance.row_count = len(day_rows)
                    instance.error_count = 0
                    instance.last_error_summary = ""
                    instance.submitted_at = None

                    csv_buffer = StringIO()
                    writer = csv.writer(csv_buffer)
                    writer.writerow(header)
                    for row in day_rows:
                        normalized = ["" if v is None else str(v) for v in row]
                        writer.writerow(normalized)

                    file_name = f"historico_{slugify(dataset_type.name) or 'dataset'}_{plant.code}_{period.isoformat()}.csv"
                    instance.raw_file.save(
                        file_name,
                        ContentFile(csv_buffer.getvalue().encode("utf-8")),
                        save=False,
                    )
                    instance.save()

                uploaded.seek(0)
                batch.source_file.save(
                    getattr(uploaded, "name", "historico") or "historico",
                    uploaded,
                    save=False,
                )
                batch.created_instances = created_count
                batch.updated_instances = updated_count
                batch.skipped_instances = skipped_count
                batch.status = HistoricalImportBatch.STATUS_DONE
                batch.finished_at = timezone.now()
                batch.save(
                    update_fields=[
                        "source_file",
                        "status",
                        "finished_at",
                        "date_column_name",
                        "total_rows",
                        "total_days",
                        "period_start",
                        "period_end",
                        "created_instances",
                        "updated_instances",
                        "skipped_instances",
                    ]
                )

                messages.success(
                    request,
                    f"Histórico importado: {created_count} creados, {updated_count} actualizados, {skipped_count} omitidos.",
                )
                record_action(
                    "UPLOAD",
                    request=request,
                    module="Ingest",
                    object_repr=f"Histórico {dataset_type.name} ({plant.code})",
                    details=f"Filas {batch.total_rows}, días {batch.total_days}",
                )
                return redirect(reverse("ingest:upload_history"))
            except Exception as exc:
                batch.status = HistoricalImportBatch.STATUS_FAILED
                batch.error_summary = str(exc)
                batch.finished_at = timezone.now()
                batch.save(update_fields=["status", "error_summary", "finished_at"])
                messages.error(request, f"Error al importar histórico: {exc}")
    else:
        form = HistoricalDatasetUploadForm(loader_plants=loader_plants)

    return render(
        request,
        "ingest/upload_historical.html",
        {
            "form": form,
            "loader_default_plant": default_plant,
        },
    )


def upload_manual(request):
    user = request.user
    if user.is_authenticated and (
        user.is_superuser
        or Membership.objects.filter(user=user, role="ADMIN", is_active=True).exists()
    ):
        messages.info(
            request,
            "Los administradores no realizan cargas directas; use el historial para revisar y gestionar cargas existentes.",
        )
        return redirect(reverse("ingest:upload_history"))

    if not user.is_authenticated:
        return redirect("login")

    loader_memberships = (
        Membership.objects.filter(user=user, role="LOADER", is_active=True)
        .select_related("plant", "project")
    )
    has_global_loader = loader_memberships.filter(
        plant__isnull=True, project__isnull=True
    ).exists()
    if has_global_loader:
        loader_plants = list(Plant.objects.values_list("id", flat=True))
        loader_projects = list(Project.objects.values_list("id", flat=True))
    else:
        loader_plants = list(
            loader_memberships.exclude(plant__isnull=True).values_list("plant_id", flat=True)
        )
        loader_projects = list(
            loader_memberships.exclude(project__isnull=True).values_list(
                "project_id", flat=True
            )
        )

    dataset_initial = request.GET.get("dataset_type")
    rows_requested = request.GET.get("rows")
    try:
        rows_extra = max(1, min(20, int(rows_requested))) if rows_requested else 1
    except ValueError:
        rows_extra = 1

    dataset_form = ManualDatasetForm(
        request.POST or None,
        loader_plants=loader_plants,
        loader_projects=loader_projects,
        initial={"dataset_type": dataset_initial} if dataset_initial else None,
    )

    selected_dataset = None
    dataset_form_valid = False
    if dataset_form.is_bound and dataset_form.is_valid():
        dataset_form_valid = True
        selected_dataset = dataset_form.cleaned_data["dataset_type"]
    elif dataset_initial and not dataset_form.is_bound:
        try:
            selected_dataset = DatasetType.objects.get(pk=dataset_initial)
        except DatasetType.DoesNotExist:
            selected_dataset = None

    row_formset = None
    columns = []
    rows_count = rows_extra
    if selected_dataset:
        ManualRowForm, columns = build_manual_row_form(selected_dataset)
        RowFormSet = formset_factory(ManualRowForm, extra=0, can_delete=True)

        initial_rows = []
        if request.method != "POST":
            period_value = dataset_form["period"].value()
            for _ in range(rows_extra):
                initial_rows.append({})
            if initial_rows and period_value:
                for column in columns:
                    if column.data_type == "DATE":
                        initial_rows[0][column.name] = period_value

        if request.method == "POST":
            row_formset = RowFormSet(request.POST, prefix="rows")
        else:
            row_formset = RowFormSet(
                prefix="rows",
                initial=initial_rows if initial_rows else None,
            )
        rows_count = row_formset.total_form_count() if row_formset else rows_extra
    else:
        row_formset = None
        rows_count = rows_extra

    selected_period = None
    if dataset_form_valid:
        selected_period = dataset_form.cleaned_data.get("period")
    else:
        selected_period = parse_date_cell(dataset_form["period"].value())

    month_lock_enabled = False
    if selected_dataset:
        month_lock_enabled = bool(
            _is_weekly_month_locked_dataset(selected_dataset)
            and month_columns_for_dataset(selected_dataset)
        )
    previous_values = {}
    if month_lock_enabled and selected_dataset and selected_period:
        previous_values = _previous_month_values(selected_dataset, selected_period)
    if month_lock_enabled and row_formset is not None:
        _apply_weekly_month_lock_to_formset(
            row_formset, selected_dataset, selected_period, previous_values
        )

    if request.method == "POST" and dataset_form_valid and row_formset is not None:
        valid = row_formset.is_valid()
        rows_data = []
        if valid:
            for form in row_formset:
                if form.cleaned_data.get("DELETE"):
                    continue
                if not form.has_changed():
                    continue
                cleaned_row = {
                    key: value
                    for key, value in form.cleaned_data.items()
                    if key != "DELETE"
                }
                rows_data.append(cleaned_row)
            if not rows_data:
                row_formset._non_form_errors = row_formset.error_class(
                    ["Debe ingresar al menos una fila con datos."]
                )
                valid = False

        if valid:
            if month_lock_enabled and previous_values:
                month_columns = month_columns_for_dataset(
                    dataset_form.cleaned_data["dataset_type"]
                )
                allowed_month = dataset_form.cleaned_data["period"].month
                for row in rows_data:
                    for col_name, month in month_columns.items():
                        if month == allowed_month:
                            continue
                        if col_name not in row or row.get(col_name) in (None, ""):
                            if month in previous_values:
                                row[col_name] = previous_values[month]

            month_error = _validate_weekly_month_rows(
                dataset_form.cleaned_data["dataset_type"],
                dataset_form.cleaned_data["period"],
                rows_data,
            )
            if month_error:
                row_formset._non_form_errors = row_formset.error_class([month_error])
                valid = False

        if valid:
            dataset_type = dataset_form.cleaned_data["dataset_type"]
            plant = dataset_form.cleaned_data["plant"]
            project = dataset_form.cleaned_data["project"]
            period = dataset_form.cleaned_data["period"]

            buffer = StringIO()
            writer = csv.writer(buffer)
            # Encabezados de plantilla manual: se usa el "name" de la columna
            header = [col.name for col in columns]
            writer.writerow(header)
            for row in rows_data:
                row_values = []
                for col in columns:
                    value = row.get(col.name)
                    if value is None:
                        row_values.append("")
                    elif col.data_type == "BOOLEAN":
                        row_values.append("1" if value else "0")
                    else:
                        row_values.append(str(value))
                writer.writerow(row_values)
            csv_content = buffer.getvalue()

            instance = DatasetInstance(
                dataset_type=dataset_type,
                plant=plant,
                project=project,
                period=period,
                state=DatasetInstance.STATE_DRAFT,
                row_count=len(rows_data),
                error_count=0,
                last_error_summary="",
            )
            if user.is_authenticated:
                membership = (
                    Membership.objects.filter(user=user, is_active=True)
                    .filter(
                        Q(plant=plant)
                        | Q(project=project)
                        | Q(plant__isnull=True, project__isnull=True)
                    )
                    .order_by("role")
                    .first()
                )
            else:
                membership = None
            instance.created_by = membership
            filename = f"manual_{dataset_type.slug}_{timezone.now():%Y%m%d%H%M%S}.csv"
            instance.raw_file.save(filename, ContentFile(csv_content), save=True)

            messages.success(
                request,
                "Datos capturados manualmente y guardados como borrador. Ahora puedes revisarlos y enviarlos a validación diaria.",
            )
            record_action(
                "UPLOAD",
                request=request,
                module="Ingest",
                object_repr=f"{instance.dataset_type.name} | {instance.period}",
                details=(
                    f"Captura manual en planta {instance.plant.code}"
                    if instance.plant
                    else f"Captura manual en proyecto {instance.project.name}"
                ),
            )
            return redirect(reverse("ingest:upload"))

    if request.user.is_authenticated:
        instances = (
            DatasetInstance.objects.select_related("dataset_type", "plant")
            .filter(created_by__user=request.user)
            .order_by("-created_at")[:10]
        )
    else:
        instances = DatasetInstance.objects.none()

    rows_more_value = min(20, rows_count + 3)
    rows_less_value = max(1, rows_count - 1)
    can_add_rows = rows_count < 20
    can_remove_rows = rows_count > 1

    default_plant = Plant.objects.filter(id=loader_plants[0]).first() if loader_plants else None
    default_project = Project.objects.filter(id=loader_projects[0]).first() if loader_projects else None

    return render(
        request,
        "ingest/manual_entry.html",
        {
            "dataset_form": dataset_form,
            "row_formset": row_formset,
            "columns": columns,
            "instances": instances,
            "rows_extra": rows_count,
            "rows_more_value": rows_more_value,
            "rows_less_value": rows_less_value,
            "can_add_rows": can_add_rows,
            "can_remove_rows": can_remove_rows,
            "selected_dataset": selected_dataset,
            "loader_default_plant": default_plant,
            "loader_default_project": default_project,
            "month_lock_enabled": month_lock_enabled,
        },
    )


def download_template(request):
    dataset_type_id = request.GET.get("dataset_type")
    if not dataset_type_id:
        return redirect("ingest:upload")

    try:
        ds = DatasetType.objects.get(
            pk=dataset_type_id,
            is_active=True,
            status=DatasetType.STATUS_APPROVED,
        )
    except DatasetType.DoesNotExist:
        raise Http404("Dataset no encontrado o no aprobado.")

    columns = ds.columns.filter(is_active=True).order_by("display_order", "name")

    if not columns.exists():
        raise Http404("El esquema seleccionado no tiene columnas activas.")

    if ds.plant:
        base_name = f"plantilla_{ds.plant.code}_{ds.name}"
    elif ds.project:
        base_name = f"plantilla_{ds.project.name}_{ds.name}"
    else:
        base_name = f"plantilla_{ds.name}"
    safe_name = f"{slugify(base_name) or 'plantilla'}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla"

    # Encabezados de plantilla descargable: se usa el "name" de la columna
    header = [col.name for col in columns]
    ws.append(header)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename=\"{safe_name}\"'

    return response


def instance_detail(request, pk):
    instance = (
        DatasetInstance.objects.select_related(
            "dataset_type", "plant", "project", "created_by__user"
        )
        .filter(pk=pk)
        .first()
    )
    if not instance:
        raise Http404("Carga no encontrada.")

    user = request.user
    if not user.is_authenticated:
        return redirect("ingest:upload_history")

    is_admin = user.is_superuser or Membership.objects.filter(
        user=user,
        role="ADMIN",
        is_active=True,
    ).exists()

    is_creator = instance.created_by and instance.created_by.user_id == user.id

    is_validator = (
        Membership.objects.filter(
            user=user,
            role="VALIDATOR",
            is_active=True,
        )
        .filter(
            Q(plant=instance.plant)
            | Q(project=instance.project)
            | Q(plant__isnull=True, project__isnull=True)
        )
        .exists()
    )

    is_loader = Membership.objects.filter(
        user=user,
        role="LOADER",
        is_active=True,
    ).filter(
        Q(plant=instance.plant)
        | Q(project=instance.project)
        | Q(plant__isnull=True, project__isnull=True)
    ).exists()

    if not (is_admin or is_creator or is_validator):
        return redirect("ingest:upload_history")

    header = []
    rows = []

    if instance.raw_file:
        file_field = instance.raw_file
        name = file_field.name or ""
        ext = name.lower().rsplit(".", 1)[-1] if "." in name else ""

        try:
            if ext in ("xlsx", "xlsm", "xltx", "xltm"):
                with file_field.open("rb") as fh:
                    wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
                    ws = wb.active
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        values = ["" if v is None else str(v) for v in row]
                        if i == 0:
                            header = values
                        else:
                            rows.append(values)
            else:
                with file_field.open("rb") as fh:
                    wrapper = TextIOWrapper(fh, encoding="utf-8", errors="ignore")
                    reader = csv.reader(wrapper)
                    for i, row in enumerate(reader):
                        values = row
                        if i == 0:
                            header = values
                        else:
                            rows.append(values)
            if rows:
                rows = rows[-100:]
                rows = list(reversed(rows))
        except Exception:
            header = []
            rows = []

    if not header and not rows:
        points = (
            PublishedDataPoint.objects.filter(instance=instance)
            .select_related("column")
            .order_by("row_index", "column__display_order", "column__name")
        )
        if points.exists():
            # Encabezados del CSV exportado: usar siempre el "name"
            header = [
                col.name
                for col in instance.dataset_type.columns.filter(is_active=True).order_by("display_order", "name")
            ]
            rows_map = {}
            for p in points:
                row = rows_map.setdefault(p.row_index, {})
                if p.numeric_value is not None:
                    value = p.numeric_value
                elif p.date_value is not None:
                    value = p.date_value.isoformat()
                elif p.bool_value is not None:
                    value = bool(p.bool_value)
                else:
                    value = p.text_value
                # Clave de columna en exportaciones: siempre el "name"
                row[p.column.name] = value
            rows = [
                [row.get(col, "") for col in header]
                for _, row in sorted(rows_map.items(), key=lambda item: item[0])
            ]
            if rows:
                rows = rows[-100:]
                rows = list(reversed(rows))

    next_url = request.GET.get("next")
    if next_url and next_url.startswith("/"):
        back_url = next_url
    elif is_loader and not is_validator and not is_admin:
        back_url = reverse("ingest:upload")
    elif is_validator and not is_admin:
        # Validador (con o sin rol de cargador): volver al detalle de validación
        back_url = reverse("validation:detail", args=[instance.pk])
    elif is_admin:
        back_url = reverse("validation:admin_overview")
    else:
        back_url = reverse("ingest:upload_history")

    return render(
        request,
        "ingest/instance_detail.html",
        {
            "instance": instance,
            "header": header,
            "rows": rows,
            "back_url": back_url,
        },
    )


def _normalize_value_for_compare(column, value):
    if value in ("", None):
        return None
    if column.data_type in ("INTEGER", "FLOAT"):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if column.data_type == "BOOLEAN":
        return bool(value)
    if column.data_type == "DATE":
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        try:
            return datetime.fromisoformat(str(value)).date()
        except (TypeError, ValueError):
            return None
    return str(value)


def _build_rows_from_points(instance_for_rows):
    points = (
        PublishedDataPoint.objects.filter(instance=instance_for_rows)
        .select_related("column")
        .order_by("row_index", "column__display_order", "column__name")
    )
    rows_map: dict[int, dict] = {}
    for p in points:
        row = rows_map.setdefault(p.row_index, {})
        if p.numeric_value is not None:
            value = p.numeric_value
        elif p.date_value is not None:
            value = p.date_value
        elif p.bool_value is not None:
            value = bool(p.bool_value)
        else:
            value = p.text_value
        row[p.column.name] = value
    return rows_map


def _extract_row_values(instance_obj, columns, row_index):
    rows = _build_rows_from_points(instance_obj)
    values = rows.get(row_index, {})
    return {column.name: values.get(column.name) for column in columns}


def _extract_initial_values(instance_obj, columns):
    return _extract_row_values(instance_obj, columns, row_index=1)


def _coerce_to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            pass
        if " " in value:
            head = value.split(" ", 1)[0]
            try:
                return datetime.fromisoformat(head).date()
            except ValueError:
                pass
    return None


def _format_value_for_display(column, value):
    if value in (None, ""):
        return "-"
    if column.data_type == "DATE":
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return str(value)
    if column.data_type == "BOOLEAN":
        return "Si" if bool(value) else "No"
    return str(value)


def _recalculate_monthly_totals(instance: DatasetInstance) -> dict:
    dataset = instance.dataset_type
    source = dataset.source_dataset
    if not source:
        return {}

    month_end = instance.period
    month_start = month_end.replace(day=1)

    source_columns = {column.name: column for column in source.columns.filter(is_active=True)}
    monthly_columns = list(dataset.columns.filter(is_active=True))
    existing_points = {
        point.column_id: point for point in PublishedDataPoint.objects.filter(instance=instance)
    }

    date_column = next((col for col in source.columns.filter(is_active=True) if col.data_type == "DATE"), None)
    if not date_column:
        return {}

    monthly_instances = list(
        DatasetInstance.objects.filter(
            dataset_type=source,
            plant=instance.plant,
            state__in=[DatasetInstance.STATE_PUBLISHED, DatasetInstance.STATE_LOCKED],
            period__gte=month_start,
            period__lte=month_end,
        ).order_by("period")
    )
    if not monthly_instances:
        return {}

    instance_rows_cache: dict[int, dict[int, dict]] = {
        daily_instance.id: _build_rows_from_points(daily_instance) for daily_instance in monthly_instances
    }

    date_entries: dict[date, dict] = {}
    for daily_instance in monthly_instances:
        rows_cache = instance_rows_cache.get(daily_instance.id, {})
        for row_index, row_values in rows_cache.items():
            raw_date = row_values.get(date_column.name) or daily_instance.period
            parsed_date = _coerce_to_date(raw_date)
            if not parsed_date or parsed_date < month_start or parsed_date > month_end:
                continue
            current = date_entries.get(parsed_date)
            should_replace = False
            if current is None:
                should_replace = True
            else:
                current_created = current["instance"].created_at
                if daily_instance.created_at > current_created:
                    should_replace = True
                elif daily_instance.id == current["instance"].id and row_index > current["row_index"]:
                    should_replace = True
            if should_replace:
                date_entries[parsed_date] = {
                    "instance": daily_instance,
                    "row_index": row_index,
                    "values": row_values,
                }

    aggregated_values: dict[str, object] = {}
    sorted_dates = sorted(date_entries.keys())

    for column in monthly_columns:
        source_column = source_columns.get(column.name)
        if not source_column:
            continue

        if column.data_type in ("INTEGER", "FLOAT"):
            total = 0.0
            has_value = False
            for entry_date in sorted_dates:
                value = date_entries[entry_date]["values"].get(column.name)
                if value in (None, ""):
                    continue
                try:
                    total += float(value)
                    has_value = True
                except (TypeError, ValueError):
                    continue
            aggregated_value = total if has_value else None
        elif column.data_type == "BOOLEAN":
            aggregated_value = any(
                bool(date_entries[entry_date]["values"].get(column.name)) for entry_date in sorted_dates
            )
        elif column.data_type == "DATE":
            aggregated_value = month_end
        else:
            aggregated_value = None
            for entry_date in sorted_dates[::-1]:
                value = date_entries[entry_date]["values"].get(column.name)
                if value not in ("", None):
                    aggregated_value = value
                    break
            if aggregated_value in ("", None):
                aggregated_value = f"Consolidado {month_start:%Y-%m}"

        point = existing_points.get(column.id)
        if not point:
            point = PublishedDataPoint(instance=instance, column=column, row_index=1)
            existing_points[column.id] = point

        point.numeric_value = None
        point.text_value = ""
        point.date_value = None
        point.bool_value = None

        if column.data_type in ("INTEGER", "FLOAT"):
            point.numeric_value = float(aggregated_value) if aggregated_value is not None else None
        elif column.data_type == "DATE":
            point.date_value = aggregated_value
        elif column.data_type == "BOOLEAN":
            point.bool_value = bool(aggregated_value)
        else:
            point.text_value = aggregated_value or ""

        point.save()
        aggregated_values[column.name] = aggregated_value

    return aggregated_values


def _submit_instance_to_validation(instance: DatasetInstance, request) -> None:
    if instance.state != DatasetInstance.STATE_DRAFT:
        return

    instance.state = DatasetInstance.STATE_SUBMITTED
    instance.last_error_summary = ""
    instance.submitted_at = timezone.now()
    instance.save(update_fields=["state", "last_error_summary", "submitted_at"])

    if (
        instance.dataset_type.is_certification
        and instance.dataset_type.validation_frequency == DatasetType.MONTHLY
    ):
        success_message = "Consolidacion enviada a validacion mensual."
        details = "Enviado a validacion mensual"
    elif instance.dataset_type.validation_frequency == DatasetType.WEEKLY:
        success_message = "Dataset enviado a validacion semanal."
        details = "Enviado a validacion semanal"
    elif instance.dataset_type.validation_frequency == DatasetType.MONTHLY:
        success_message = "Dataset enviado a validacion mensual."
        details = "Enviado a validacion mensual"
    elif instance.dataset_type.validation_frequency == DatasetType.FLEXIBLE:
        success_message = "Dataset de proyecciones enviado a validacion."
        details = "Enviado a validacion (proyecciones)"
    else:
        success_message = "Dataset enviado a validacion diaria."
        details = "Enviado a validacion diaria"

    messages.success(request, success_message)
    record_action(
        "SUBMIT",
        request=request,
        module="Ingest",
        object_repr=f"{instance.dataset_type.name} | {instance.period}",
        details=details,
    )


def certification_review(request, pk):
    instance = (
        DatasetInstance.objects.select_related(
            "dataset_type", "plant", "project", "created_by__user"
        )
        .filter(pk=pk)
        .first()
    )
    if not instance:
        raise Http404("Carga no encontrada.")

    dataset = instance.dataset_type
    if not (dataset.is_certification and dataset.validation_frequency == DatasetType.MONTHLY):
        raise Http404("Esta vista solo aplica a consolidaciones mensuales.")

    if not request.user.is_authenticated:
        return redirect("login")

    is_admin = request.user.is_superuser or Membership.objects.filter(
        user=request.user,
        role="ADMIN",
        is_active=True,
    ).exists()

    loader_membership = (
        Membership.objects.filter(
            user=request.user,
            role="LOADER",
            is_active=True,
        )
        .filter(
            Q(plant=instance.plant)
            | Q(project=instance.project)
            | Q(plant__isnull=True, project__isnull=True)
        )
        .first()
    )

    if not loader_membership and not is_admin:
        raise Http404("No tienes permisos para revisar esta consolidación.")

    can_edit = loader_membership is not None and instance.state == DatasetInstance.STATE_DRAFT

    form_type = request.POST.get("form_type") if request.method == "POST" else None

    has_month_override = instance.change_requests.filter(
        target_period=instance.period,
        target_instance__isnull=True,
    ).exists()
    if dataset.source_dataset and not has_month_override:
        _recalculate_monthly_totals(instance)

    ManualRowForm, columns = build_manual_row_form(dataset)
    existing_points = (
        PublishedDataPoint.objects.filter(instance=instance)
        .select_related("column")
    )
    existing_values = {}
    for point in existing_points:
        col = point.column
        if col.data_type in ("INTEGER", "FLOAT"):
            value = point.numeric_value
        elif col.data_type == "DATE":
            value = point.date_value
        elif col.data_type == "BOOLEAN":
            value = point.bool_value
        else:
            value = point.text_value
        existing_values[col.name] = value

    initial_data = {col.name: existing_values.get(col.name) for col in columns}
    bind_month_post = request.method == "POST" and form_type == "month" and can_edit
    row_form = ManualRowForm(
        request.POST if bind_month_post else None,
        prefix="row",
        initial=initial_data,
    )
    justification_form = CertificationJustificationForm(
        request.POST if bind_month_post else None,
        prefix="justification",
    )
    justification_field = justification_form.fields["justification"]
    justification_field.widget.attrs["data-justification-input"] = "true"

    justification_value = ""
    if justification_form.is_bound:
        justification_value = justification_form.data.get("justification-justification", "").strip()
    justification_ready = bool(justification_value)

    for field in row_form.fields.values():
        css_class = field.widget.attrs.get("class", "")
        field.widget.attrs["class"] = (css_class + " js-cert-field").strip()
        field.widget.attrs["data-requires-justification"] = "true"
        if can_edit and not justification_ready:
            field.widget.attrs["disabled"] = True

    if not can_edit:
        for field in row_form.fields.values():
            field.disabled = True
        justification_field = justification_form.fields["justification"]
        justification_field.widget.attrs["readonly"] = True
        justification_field.widget.attrs["disabled"] = True
        justification_field.disabled = True

    attachment_errors: list[str] = []
    files_to_process = []
    submit_action = request.POST.get("submit_action", "save")

    if bind_month_post and can_edit:
        files_to_process = request.FILES.getlist("support_files")
        for uploaded in files_to_process:
            content_type = (uploaded.content_type or "").lower()
            if content_type and not content_type.startswith("image/"):
                attachment_errors.append(f"{uploaded.name}: formato no permitido.")

        if row_form.is_valid() and justification_form.is_valid() and not attachment_errors:
            justification_text = justification_form.cleaned_data["justification"].strip()
            if not justification_text:
                justification_form.add_error(
                    "justification", "Debes ingresar una justificación para los cambios."
                )
            else:
                cleaned = row_form.cleaned_data
                changed_columns = []
                for column in columns:
                    new_value = cleaned.get(column.name)
                    old_value = existing_values.get(column.name)
                    if _normalize_value_for_compare(column, new_value) != _normalize_value_for_compare(
                        column, old_value
                    ):
                        changed_columns.append((column, new_value))

                if not changed_columns:
                    if submit_action == "send":
                        _submit_instance_to_validation(instance, request)
                        return redirect("ingest:upload_history")
                    row_form.add_error(
                        None, "No se detectaron cambios en los valores consolidados."
                    )
                else:
                    for column, new_value in changed_columns:
                        point, _ = PublishedDataPoint.objects.get_or_create(
                            instance=instance,
                            column=column,
                            row_index=1,
                        )
                        point.numeric_value = None
                        point.text_value = ""
                        point.date_value = None
                        point.bool_value = None

                        if column.data_type in ("INTEGER", "FLOAT"):
                            point.numeric_value = float(new_value) if new_value not in (None, "") else None
                        elif column.data_type == "DATE":
                            if isinstance(new_value, datetime):
                                point.date_value = new_value.date()
                            else:
                                point.date_value = new_value or None
                        elif column.data_type == "BOOLEAN":
                            point.bool_value = bool(new_value)
                        else:
                            point.text_value = new_value or ""
                        point.save()

                    change_request = DatasetChangeRequest.objects.create(
                        instance=instance,
                        submitted_by=loader_membership,
                        justification=justification_text,
                        target_period=instance.period,
                    )
                    for uploaded in files_to_process:
                        DatasetChangeAttachment.objects.create(
                            request=change_request,
                            file=uploaded,
                            original_name=uploaded.name or "",
                        )

                    if submit_action == "send":
                        _submit_instance_to_validation(instance, request)
                        return redirect("ingest:upload_history")
                    messages.success(request, "Cambios registrados y justificados correctamente.")
                    return redirect("ingest:certification_review", pk=instance.pk)

    column_fields = [(column, row_form[column.name]) for column in columns]

    change_requests_qs = list(
        instance.change_requests.select_related("submitted_by__user")
        .prefetch_related("attachments")
        .order_by("-created_at")
    )
    changed_daily_periods = {
        req.target_period
        for req in change_requests_qs
        if req.target_period
        and req.target_period != instance.period
    }

    daily_forms = []
    daily_rows = []
    daily_columns = []
    daily_display_columns = []
    open_daily_modal_id = None
    if dataset.source_dataset:
        DailyRowForm, daily_columns = build_manual_row_form(dataset.source_dataset)
        daily_display_columns = [col for col in daily_columns if col.data_type != "DATE"]
        if not daily_display_columns:
            daily_display_columns = list(daily_columns)
        month_start = instance.period.replace(day=1)
        month_year = month_start.year
        month_number = month_start.month
        month_last_day = calendar.monthrange(month_year, month_number)[1]
        month_end = date(month_year, month_number, month_last_day)
        month_dates = [date(month_year, month_number, day) for day in range(1, month_last_day + 1)]
        monthly_instances = list(
            DatasetInstance.objects.filter(
                dataset_type=dataset.source_dataset,
                plant=instance.plant,
                state__in=[DatasetInstance.STATE_PUBLISHED, DatasetInstance.STATE_LOCKED],
                period__gte=month_start,
                period__lte=month_end,
            ).order_by("period")
        )
        instance_rows_cache: dict[int, dict[int, dict]] = {
            daily_instance.id: _build_rows_from_points(daily_instance)
            for daily_instance in monthly_instances
        }
        date_column = next((col for col in daily_columns if col.data_type == "DATE"), None)
        date_entries: dict[date, dict] = {}
        for daily_instance in monthly_instances:
            rows_cache = instance_rows_cache.get(daily_instance.id, {})
            for row_index, row_values in rows_cache.items():
                if date_column:
                    raw_date = row_values.get(date_column.name)
                else:
                    raw_date = daily_instance.period
                parsed_date = _coerce_to_date(raw_date)
                if not parsed_date:
                    continue
                if parsed_date < month_start or parsed_date > month_end:
                    continue
                current = date_entries.get(parsed_date)
                should_replace = False
                if current is None:
                    should_replace = True
                else:
                    current_created = current["instance"].created_at
                    if daily_instance.created_at > current_created:
                        should_replace = True
                    elif daily_instance.id == current["instance"].id and row_index > current["row_index"]:
                        should_replace = True
                if should_replace:
                    date_entries[parsed_date] = {
                        "instance": daily_instance,
                        "row_index": row_index,
                        "values": row_values,
                    }

        daily_target_id = request.POST.get("day_instance_id") if form_type == "daily" else None
        daily_target_row = request.POST.get("day_row_index") if form_type == "daily" else None

        for target_date in month_dates:
            entry = date_entries.get(target_date)
            if entry:
                daily_instance = entry["instance"]
                row_index = entry["row_index"]
                initial_values = {
                    column.name: entry["values"].get(column.name) for column in daily_columns
                }
                for column in daily_columns:
                    if column.data_type == "DATE":
                        normalized = _coerce_to_date(initial_values.get(column.name))
                        if normalized:
                            initial_values[column.name] = normalized
                        else:
                            initial_values[column.name] = target_date
                prefix = f"day-{daily_instance.id}-{row_index}"
                bind_daily = (
                    request.method == "POST"
                    and form_type == "daily"
                    and daily_target_id == str(daily_instance.id)
                    and daily_target_row == str(row_index)
                    and can_edit
                )
                day_form = DailyRowForm(
                    request.POST if bind_daily else None,
                    prefix=prefix,
                    initial=initial_values,
                )
                justification_value = (
                    request.POST.get("day_justification", "").strip() if bind_daily else ""
                )
                attachment_errors_row: list[str] = []
                justification_error = False

                for field in day_form.fields.values():
                    css_class = field.widget.attrs.get("class", "")
                    field.widget.attrs["class"] = (css_class + " js-cert-field").strip()
                    field.widget.attrs["data-requires-justification"] = "true"
                    field.widget.attrs["disabled"] = True

                if bind_daily:
                    files = request.FILES.getlist("day_support_files")
                    for uploaded in files:
                        content_type = (uploaded.content_type or "").lower()
                        if content_type and not content_type.startswith("image/"):
                            attachment_errors_row.append(f"{uploaded.name}: formato no permitido.")

                    if day_form.is_valid() and justification_value and not attachment_errors_row:
                        cleaned = day_form.cleaned_data
                        changed_columns = []
                        for column in daily_columns:
                            new_value = cleaned.get(column.name)
                            old_value = initial_values.get(column.name)
                            if _normalize_value_for_compare(column, new_value) != _normalize_value_for_compare(
                                column, old_value
                            ):
                                changed_columns.append((column, new_value))

                        if not changed_columns:
                            day_form.add_error(None, "No se detectaron cambios en este día.")
                        else:
                            for column, new_value in changed_columns:
                                point, _ = PublishedDataPoint.objects.get_or_create(
                                    instance=daily_instance,
                                    column=column,
                                    row_index=row_index,
                                )
                                point.numeric_value = None
                                point.text_value = ""
                                point.date_value = None
                                point.bool_value = None

                                if column.data_type in ("INTEGER", "FLOAT"):
                                    point.numeric_value = float(new_value) if new_value not in (None, "") else None
                                elif column.data_type == "DATE":
                                    if isinstance(new_value, datetime):
                                        point.date_value = new_value.date()
                                    else:
                                        point.date_value = new_value or None
                                elif column.data_type == "BOOLEAN":
                                    point.bool_value = bool(new_value)
                                else:
                                    point.text_value = new_value or ""
                                point.save()

                            change_request = DatasetChangeRequest.objects.create(
                                instance=instance,
                                submitted_by=loader_membership,
                                justification=f"Día {target_date}: {justification_value}",
                                target_instance=daily_instance,
                                target_period=target_date,
                            )
                            for uploaded in files:
                                DatasetChangeAttachment.objects.create(
                                    request=change_request,
                                    file=uploaded,
                                    original_name=uploaded.name or "",
                                )

                            _recalculate_monthly_totals(instance)
                            messages.success(request, f"Datos del día {target_date} actualizados.")
                            return redirect("ingest:certification_review", pk=instance.pk)

                    if not justification_value:
                        justification_error = True

                fields = [(column, day_form[column.name]) for column in daily_columns]
                form_id = f"daily-form-{daily_instance.id}-{row_index}"
                form_key = f"{daily_instance.id}-{row_index}"
                auto_open_modal = bind_daily and (
                    not day_form.is_valid() or attachment_errors_row or justification_error
                )
                if auto_open_modal:
                    open_daily_modal_id = form_key
                daily_forms.append(
                    {
                        "instance": daily_instance,
                        "row_index": row_index,
                        "form": day_form,
                        "fields": fields,
                        "form_id": form_id,
                        "form_key": form_key,
                        "attachment_errors": attachment_errors_row if attachment_errors_row else None,
                        "justification_value": justification_value,
                        "justification_error": justification_error,
                        "justification_ready": bind_daily and bool(justification_value),
                        "auto_open": auto_open_modal,
                        "date": target_date,
                    }
                )

                display_values = [
                    {
                        "column": column,
                        "display": _format_value_for_display(column, initial_values.get(column.name)),
                    }
                    for column in daily_display_columns
                ]
                daily_rows.append(
                    {
                        "instance": daily_instance,
                        "row_index": row_index,
                        "date": target_date,
                        "values": display_values,
                        "has_instance": True,
                        "form_key": form_key,
                        "was_changed": target_date in changed_daily_periods,
                    }
                )
            else:
                display_values = [
                    {
                        "column": column,
                        "display": "-",
                    }
                    for column in daily_display_columns
                ]
                daily_rows.append(
                    {
                        "instance": None,
                        "date": target_date,
                        "values": display_values,
                        "has_instance": False,
                        "form_key": None,
                        "was_changed": False,
                    }
                )

    change_requests = change_requests_qs

    attachment_errors_display = attachment_errors if attachment_errors else None

    return render(
        request,
        "ingest/certification_review.html",
        {
            "instance": instance,
            "form": row_form,
            "columns": columns,
            "column_fields": column_fields,
            "daily_forms": daily_forms,
            "daily_rows": daily_rows,
            "daily_columns": daily_columns,
            "daily_display_columns": daily_display_columns,
            "can_edit": can_edit,
            "justification_form": justification_form,
            "justification_ready": justification_ready,
            "attachment_errors": attachment_errors_display,
            "change_requests": change_requests,
            "open_daily_modal_id": open_daily_modal_id,
        },
    )


def upload_history(request):
    loader_certifications_pending = []
    loader_certifications_rejected = []
    loader_certifications_history = []
    loader_memberships = []
    has_global_loader = False
    loader_plants = []
    loader_projects = []

    if request.user.is_authenticated:
        loader_memberships = list(
            Membership.objects.filter(
                user=request.user,
                role="LOADER",
                is_active=True,
            ).select_related("plant", "project")
        )
        has_global_loader = any(
            m.plant_id is None and m.project_id is None for m in loader_memberships
        )
        loader_plants = [m.plant_id for m in loader_memberships if m.plant_id]
        loader_projects = [m.project_id for m in loader_memberships if m.project_id]

        queryset = DatasetInstance.objects.select_related(
            "dataset_type", "plant", "project", "created_by__user"
        )

        if request.user.is_superuser or Membership.objects.filter(
            user=request.user,
            role="ADMIN",
            is_active=True,
        ).exists():
            instances = queryset.order_by("-created_at")[:100]
        else:
            instances = queryset.filter(created_by__user=request.user).order_by("-created_at")[:50]
    else:
        instances = DatasetInstance.objects.none()

    if request.user.is_authenticated and loader_memberships:
        profile = getattr(request.user, "profile", None)
        if profile:
            profile.last_seen_validation_status = timezone.now()
            profile.save(update_fields=["last_seen_validation_status"])

        _, prev_month_end = previous_month_range()
        base_cert_qs = DatasetInstance.objects.select_related("dataset_type", "plant").filter(
            dataset_type__validation_frequency=DatasetType.MONTHLY,
            dataset_type__is_certification=True,
            period=prev_month_end,
            state=DatasetInstance.STATE_DRAFT,
        )
        if not has_global_loader:
            if loader_plants:
                base_cert_qs = base_cert_qs.filter(plant_id__in=loader_plants)
            else:
                base_cert_qs = base_cert_qs.none()
        loader_certifications_pending = list(
            base_cert_qs.filter(last_error_summary="").order_by("plant__code", "dataset_type__name")
        )
        loader_certifications_rejected = list(
            base_cert_qs.filter(last_error_summary__gt="").order_by("plant__code", "dataset_type__name")
        )
        if loader_certifications_pending and profile:
            profile.last_seen_certification_alert = timezone.now()
            profile.save(update_fields=["last_seen_certification_alert"])

        history_qs = (
            DatasetInstance.objects.select_related("dataset_type", "plant")
            .filter(
                dataset_type__validation_frequency=DatasetType.MONTHLY,
                dataset_type__is_certification=True,
            )
            .order_by("-updated_at")
        )
        if not has_global_loader:
            if loader_plants:
                history_qs = history_qs.filter(plant_id__in=loader_plants)
            else:
                history_qs = history_qs.none()
        loader_certifications_history = list(history_qs[:10])

    historical_batches = []
    if request.user.is_authenticated and loader_memberships:
        batch_qs = HistoricalImportBatch.objects.select_related("dataset_type", "plant").order_by(
            "-created_at"
        )
        if not has_global_loader:
            if loader_plants:
                batch_qs = batch_qs.filter(plant_id__in=loader_plants)
            else:
                batch_qs = batch_qs.none()

        batch_qs = batch_qs[:10]
        for batch in batch_qs:
            inst_qs = DatasetInstance.objects.filter(historical_batch=batch)
            counts = inst_qs.aggregate(
                draft=Count("id", filter=Q(state=DatasetInstance.STATE_DRAFT)),
                submitted=Count("id", filter=Q(state=DatasetInstance.STATE_SUBMITTED)),
                published=Count("id", filter=Q(state=DatasetInstance.STATE_PUBLISHED)),
                start=Min("period"),
                end=Max("period"),
            )
            historical_batches.append(
                {
                    "batch": batch,
                    "draft_count": counts["draft"] or 0,
                    "submitted_count": counts["submitted"] or 0,
                    "published_count": counts["published"] or 0,
                    "period_start": batch.period_start or counts["start"],
                    "period_end": batch.period_end or counts["end"],
                }
            )

    is_validator = (
        Membership.objects.filter(
            user=request.user,
            role="VALIDATOR",
            is_active=True,
        ).exists()
        if request.user.is_authenticated
        else False
    )
    is_admin = (
        request.user.is_authenticated
        and (
            request.user.is_superuser
            or Membership.objects.filter(
                user=request.user,
                role="ADMIN",
                is_active=True,
            ).exists()
        )
    )

    return render(
        request,
        "ingest/upload_history.html",
        {
            "instances": instances,
            "is_validator": is_validator,
            "is_admin": is_admin,
            "loader_certifications_pending": loader_certifications_pending,
            "loader_certifications_rejected": loader_certifications_rejected,
            "loader_certifications_history": loader_certifications_history,
            "historical_batches": historical_batches,
        },
    )


def submit_instance(request, pk):
    if request.method != "POST":
        return redirect("ingest:upload_history")

    instance = (
        DatasetInstance.objects.select_related(
            "dataset_type", "plant", "project", "created_by__user"
        )
        .filter(pk=pk)
        .first()
    )
    if not instance:
        raise Http404("Carga no encontrada.")

    user = request.user
    if not user.is_authenticated:
        return redirect("ingest:upload_history")

    can_submit = instance.created_by and instance.created_by.user_id == user.id
    if (
        not can_submit
        and instance.dataset_type.is_certification
        and Membership.objects.filter(
            user=user,
            role="LOADER",
            is_active=True,
        )
        .filter(
            Q(plant=instance.plant)
            | Q(project=instance.project)
            | Q(plant__isnull=True, project__isnull=True)
        )
        .exists()
    ):
        can_submit = True

    if not can_submit:
        return redirect("ingest:upload_history")

    if instance.state == DatasetInstance.STATE_DRAFT:
        _submit_instance_to_validation(instance, request)

    return redirect("ingest:upload")


def submit_historical_batch(request, batch_id: int):
    if request.method != "POST":
        return redirect("ingest:upload_history")

    user = request.user
    if not user.is_authenticated:
        return redirect("login")

    batch = (
        HistoricalImportBatch.objects.select_related("dataset_type", "plant")
        .filter(pk=batch_id)
        .first()
    )
    if not batch:
        raise Http404("Importación histórica no encontrada.")

    is_loader = Membership.objects.filter(user=user, role="LOADER", is_active=True).exists()
    if not is_loader:
        return redirect("ingest:upload_history")

    can_access_plant = Membership.objects.filter(
        user=user, role="LOADER", is_active=True
    ).filter(
        Q(plant=batch.plant)
        | Q(plant__isnull=True, project__isnull=True)
    ).exists()
    if not can_access_plant:
        return redirect("ingest:upload_history")

    now = timezone.now()
    updated = DatasetInstance.objects.filter(
        historical_batch=batch,
        state=DatasetInstance.STATE_DRAFT,
    ).update(
        state=DatasetInstance.STATE_SUBMITTED,
        last_error_summary="",
        submitted_at=now,
        updated_at=now,
    )

    if updated:
        messages.success(request, f"Histórico enviado a validación: {updated} días.")
        record_action(
            "SUBMIT",
            request=request,
            module="Ingest",
            object_repr=f"Histórico {batch.dataset_type.name} | {batch.plant.code}",
            details=f"Envío masivo ({updated} instancias)",
        )
    else:
        messages.info(request, "No hay instancias en borrador para enviar en este histórico.")

    return redirect("ingest:upload_history")


def edit_instance(request, pk):
    instance = (
        DatasetInstance.objects.select_related("dataset_type", "plant", "created_by__user")
        .filter(pk=pk)
        .first()
    )
    if not instance:
        raise Http404("Carga no encontrada.")

    user = request.user
    if not user.is_authenticated or not instance.created_by or instance.created_by.user_id != user.id:
        return redirect("ingest:upload_history")

    if instance.state != DatasetInstance.STATE_DRAFT:
        messages.info(request, "Solo se pueden editar datasets en estado borrador.")
        return redirect("ingest:upload_history")

    if request.method == "POST":
        form = DatasetInstanceEditForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            upload_error = _validate_weekly_month_file(
                instance.dataset_type,
                instance.period,
                form.cleaned_data.get("raw_file"),
            )
            if upload_error:
                form.add_error("raw_file", upload_error)
            else:
                instance = form.save(commit=False)
                instance.row_count = 0
                instance.error_count = 0
                instance.last_error_summary = ""
                instance.state = DatasetInstance.STATE_DRAFT
                instance.save()
                justification = (form.cleaned_data.get("justification") or "").strip()
                details = "Archivo corregido y reemplazado"
                if instance.dataset_type.is_one_time:
                    details = f"Actualizacion de carga unica. Justificacion: {justification}"
                messages.success(
                    request,
                    "Dataset actualizado. Ahora puedes enviarlo a validacion diaria.",
                )
                record_action(
                    "EDIT",
                    request=request,
                    module="Ingest",
                    object_repr=f"{instance.dataset_type.name} | {instance.period}",
                    details=details,
                )
                return redirect("ingest:upload")
    else:
        form = DatasetInstanceEditForm(instance=instance)

    return render(
        request,
        "ingest/instance_edit.html",
        {
            "instance": instance,
            "form": form,
        },
    )


def delete_instance(request, pk):
    if request.method != "POST":
        return redirect("ingest:upload_history")

    instance = DatasetInstance.objects.select_related("created_by__user").filter(pk=pk).first()
    if not instance:
        raise Http404("Carga no encontrada.")

    user = request.user
    if not user.is_authenticated or not instance.created_by or instance.created_by.user_id != user.id:
        return redirect("ingest:upload_history")

    if instance.state != DatasetInstance.STATE_DRAFT:
        messages.info(request, "Solo se pueden eliminar datasets en estado borrador.")
        return redirect("ingest:upload_history")

    record_action(
        "DELETE",
        request=request,
        module="Ingest",
        object_repr=f"{instance.dataset_type.name} | {instance.period}",
        details="Carga eliminada",
    )
    instance.delete()
    messages.success(request, "Dataset eliminado correctamente.")
    return redirect("ingest:upload")
